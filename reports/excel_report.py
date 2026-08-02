"""
Excel Report Exporter for CICMS using openpyxl.
Generates multi-tab, professionally formatted Excel workbooks with KPI summaries and detailed registers.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from typing import Dict, Any
from pathlib import Path
from database.connection import get_db

def generate_excel_report(output_path: str) -> str:
    """Generates a complete CICMS Excel Master Report."""
    db = get_db()
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------------------
    # Styles
    # -------------------------------------------------------------------------
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    kpi_title_font = Font(name="Calibri", size=12, bold=True, color="1F497D")
    kpi_val_font = Font(name="Calibri", size=14, bold=True, color="002060")
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    kpi_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # -------------------------------------------------------------------------
    # Sheet 1: Executive KPI Dashboard Summary
    # -------------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    ws_summary["A1"] = "CRIMINAL INVESTIGATION AND CASE MANAGEMENT SYSTEM"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = "Master Executive KPI & Performance Report"
    ws_summary["A2"].font = Font(name="Calibri", size=12, italic=True, color="595959")
    
    # Fetch KPIs
    kpi_data = [
        ("Total FIRs Registered", "SELECT COUNT(*) FROM firs"),
        ("Total Active Cases", "SELECT COUNT(*) FROM cases WHERE status IN ('Open', 'In Progress')"),
        ("Total Closed Cases", "SELECT COUNT(*) FROM cases WHERE status = 'Closed'"),
        ("Total Evidence Items Vaulted", "SELECT COUNT(*) FROM evidence"),
        ("Court Cases Pending", "SELECT COUNT(*) FROM court_cases WHERE verdict = 'Pending'"),
        ("Court Convictions", "SELECT COUNT(*) FROM court_cases WHERE verdict = 'Convicted'")
    ]
    
    row_idx = 5
    for title, query in kpi_data:
        val_res = db.fetch_one(query)
        val = list(val_res.values())[0] if val_res else 0
        
        ws_summary.cell(row=row_idx, column=1, value=title).font = kpi_title_font
        ws_summary.cell(row=row_idx, column=1).fill = kpi_fill
        
        cell_val = ws_summary.cell(row=row_idx, column=2, value=val)
        cell_val.font = kpi_val_font
        cell_val.fill = kpi_fill
        cell_val.alignment = Alignment(horizontal="right")
        
        ws_summary.cell(row=row_idx, column=1).border = thin_border
        ws_summary.cell(row=row_idx, column=2).border = thin_border
        row_idx += 1

    # -------------------------------------------------------------------------
    # Sheet 2: FIR Register
    # -------------------------------------------------------------------------
    ws_fir = wb.create_sheet(title="FIR Register")
    firs = db.fetch_all("SELECT fir_number, complainant_name, incident_date, incident_location, status FROM firs LIMIT 500")
    if firs:
        headers = ["FIR Number", "Complainant Name", "Incident Date", "Incident Location", "Status"]
        ws_fir.append(headers)
        for col in range(1, 6):
            cell = ws_fir.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        for f in firs:
            ws_fir.append([f["fir_number"], f["complainant_name"], str(f["incident_date"]), f["incident_location"], f["status"]])

    # Adjust Column Widths
    for ws in [ws_summary, ws_fir]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output_path)
    return output_path
